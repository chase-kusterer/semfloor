"""
Keyword import/export for the setup page.

Two formats are understood on upload (auto-detected from the header row):

1. **Native** — the same CSV this app downloads. Columns:
   Keyword, Asset class, Volume, Conversion rate, Order value, Reserve price, Notes
   Round-trips perfectly: download, edit in Excel/Sheets, upload again.

2. **Google Keyword Planner** — the "Keyword ideas" export. Relevant columns:
   Keyword, Avg. monthly searches, Competition, Top of page bid (low range),
   Top of page bid (high range). GKP has no conversion rate / order value /
   reserve, so we map:
     - searches -> volume (ranges like "1K – 10K" become the midpoint)
     - low top-of-page bid -> reserve price
     - Competition -> a default conversion rate (High 4%, Medium 2.5%, Low 1.2%)
     - order value derived as high bid / conversion rate, so a keyword's fair
       value (conv x order value) in the auction matches its real-world bid range
     - Three month change / YoY change -> notes (flavor), everything else ignored

GKP's real CSV export is UTF-16, tab-separated, with two metadata lines above
the header — all handled here. .xlsx uploads are also accepted (both formats).
"""
from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation

from .models import Keyword

# ---------------------------------------------------------------------------
# Asset classes: the label vocabulary used by the starter pack, with the
# definitions shown on the setup page. `asset_class` stays free text — these
# are documentation, not an enum.
# ---------------------------------------------------------------------------
ASSET_CLASS_DEFINITIONS = [
    ("Branded blue chip", "Your own brand name. Cheap clicks, very high conversion — reliable profit, but limited volume."),
    ("High-intent niche", "Low volume, but searchers are ready to act, so conversion is strong. Often underpriced."),
    ("B2B heavyweight", "Expensive clicks and modest conversion, but each order is worth a lot. High stakes per click."),
    ("Mid-cap", "Moderate volume, decent conversion, fair prices. A balanced core holding."),
    ("Crowded momentum", "Popular term everyone bids on. Volume is there, but competition pushes prices toward break-even."),
    ("High-volume, volatile", "Huge, ambiguous audience. Tons of clicks, weak intent — profitable only at low prices."),
    ("Commodity churn", "Price-shopper territory. Big volume, thin margins, low order values."),
    ("Speculative / ambiguous", "The searcher could want anything. Cheap, high-volume, low conversion — a lottery ticket."),
]

# Default conversion rate by GKP Competition tier.
_COMPETITION_CVR = {"high": 0.04, "medium": 0.025, "low": 0.012}
_DEFAULT_CVR = 0.02

# Native export/import header (also what the download produces).
NATIVE_HEADER = ["Keyword", "Asset class", "Volume", "Conversion rate",
                 "Order value", "Reserve price", "Notes"]


# --- small parsing helpers --------------------------------------------------

def _num(raw, default=None):
    """'12,000' / '1.5' / '$3.50' -> float; returns default if unparseable."""
    if raw is None:
        return default
    s = re.sub(r"[^0-9.\-]", "", str(raw))
    if not s or s in {".", "-"}:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _volume(raw, default=5000):
    """GKP volumes can be '12,000' or a range like '1K – 10K'. Use the midpoint."""
    if raw is None:
        return default
    s = str(raw).strip().lower().replace(",", "")
    parts = re.split(r"[–—-]| to ", s)  # range separators GKP uses
    vals = []
    for p in parts:
        m = re.match(r"\s*([0-9.]+)\s*([km]?)", p.strip())
        if m and m.group(1):
            v = float(m.group(1)) * {"": 1, "k": 1_000, "m": 1_000_000}[m.group(2)]
            vals.append(v)
    if not vals:
        return default
    return max(1, int(sum(vals) / len(vals)))


def _dec(val, default):
    try:
        return Decimal(str(round(float(val), 2)))
    except (TypeError, ValueError, InvalidOperation):
        return Decimal(default)


# --- reading files ----------------------------------------------------------

def _decode(data: bytes) -> str:
    """GKP exports are UTF-16; everything else is (nearly always) UTF-8."""
    if data[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return data.decode("utf-16")
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("utf-16")


def _rows_from_csv(text: str) -> list[list[str]]:
    # Sniff the delimiter from the first non-empty lines (GKP uses tabs).
    sample = "\n".join(line for line in text.splitlines() if line.strip())[:2000]
    delimiter = "\t" if sample.count("\t") >= sample.count(",") and "\t" in sample else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    return [["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)]


def _header_index(rows: list[list[str]]):
    """Find the header row (GKP puts 1-2 metadata lines above it)."""
    for i, row in enumerate(rows[:10]):
        cells = [c.strip().lower() for c in row]
        if "keyword" in cells:
            return i
    return None


# --- the public API ---------------------------------------------------------

class KeywordImportError(Exception):
    pass


def parse_keyword_upload(filename: str, data: bytes) -> list[dict]:
    """
    Parse an uploaded CSV/XLSX (native or Google Keyword Planner format) into a
    list of dicts with the Keyword model's field names. Raises KeywordImportError
    with a friendly message on anything unusable.
    """
    if not data:
        raise KeywordImportError("The uploaded file is empty.")
    if filename.lower().endswith((".xlsx", ".xlsm")):
        try:
            rows = _rows_from_xlsx(data)
        except Exception:
            raise KeywordImportError("Couldn't read that .xlsx file — is it a real Excel workbook?")
    else:
        rows = _rows_from_csv(_decode(data))

    hi = _header_index(rows)
    if hi is None:
        raise KeywordImportError(
            'No "Keyword" column found. Upload a Google Keyword Planner export '
            "or a CSV downloaded from this page.")
    header = [c.strip().lower() for c in rows[hi]]
    col = {name: idx for idx, name in enumerate(header)}

    def get(row, *names):
        for n in names:
            if n in col and col[n] < len(row):
                return row[col[n]]
        return None

    is_native = "conversion rate" in col
    parsed = []
    for row in rows[hi + 1:]:
        label = (get(row, "keyword") or "").strip()
        if not label:
            continue
        if is_native:
            cvr = _num(get(row, "conversion rate"), 0.03)
            cvr = cvr / 100 if cvr and cvr > 1 else cvr  # accept 3 or 0.03
            parsed.append({
                "label": label,
                "asset_class": (get(row, "asset class") or "").strip(),
                "search_volume": _volume(get(row, "volume", "search volume", "avg. monthly searches")),
                "conversion_rate": round(cvr or 0.03, 4),
                "order_value": _dec(_num(get(row, "order value"), 50), "50.00"),
                "reserve_price": _dec(_num(get(row, "reserve price"), 0.5), "0.50"),
                "notes": (get(row, "notes") or "").strip(),
            })
        else:
            competition = (get(row, "competition") or "").strip().lower()
            cvr = _COMPETITION_CVR.get(competition, _DEFAULT_CVR)
            low = _num(get(row, "top of page bid (low range)"), None)
            high = _num(get(row, "top of page bid (high range)"), None)
            # Fair value (cvr x order value) ~= the real high top-of-page bid.
            order_value = _dec((high / cvr) if high else 50, "50.00")
            reserve = _dec(max(low, 0.10) if low else 0.50, "0.50")
            trend_bits = []
            for name, tag in [("three month change", "3-mo"), ("yoy change", "YoY")]:
                v = (get(row, name) or "").strip()
                if v:
                    trend_bits.append(f"{tag} {v}")
            parsed.append({
                "label": label,
                "asset_class": f"{competition.capitalize()} competition" if competition else "",
                "search_volume": _volume(get(row, "avg. monthly searches")),
                "conversion_rate": cvr,
                "order_value": order_value,
                "reserve_price": reserve,
                "notes": ("Keyword Planner import" + (": " + ", ".join(trend_bits) if trend_bits else "")),
            })
    if not parsed:
        raise KeywordImportError("Found the header but no keyword rows underneath it.")
    return parsed


def import_keywords(game, parsed: list[dict], replace: bool = False):
    """
    Write parsed rows into the game. `replace=False` appends/updates by label;
    `replace=True` deletes existing keywords (and pending rounds) first.
    Returns (created, updated).
    """
    if replace:
        game.rounds.all().delete()
        game.keywords.all().delete()
    existing = {k.label: k for k in game.keywords.all()}
    order = (game.keywords.order_by("-order").values_list("order", flat=True).first() or 0)
    created = updated = 0
    for row in parsed:
        kw = existing.get(row["label"])
        if kw:
            for field in ("asset_class", "search_volume", "conversion_rate",
                          "order_value", "reserve_price", "notes"):
                setattr(kw, field, row[field])
            kw.save()
            updated += 1
        else:
            order += 1
            Keyword.objects.create(game=game, order=order, **row)
            created += 1
    return created, updated


def export_keywords_csv(game) -> str:
    """The native CSV: edit it in Excel/Sheets and upload it back."""
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(NATIVE_HEADER)
    for k in game.keywords.all():
        w.writerow([k.label, k.asset_class, k.search_volume, k.conversion_rate,
                    k.order_value, k.reserve_price, k.notes])
    return out.getvalue()
